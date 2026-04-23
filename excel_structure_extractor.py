# -*- coding: utf-8 -*-
"""
Excel 구조 추출기 (Python 3.10+)

필수 패키지:
- openpyxl

선택 패키지:
- xlwings (VBA 모듈 추출용, 없어도 기본 기능 동작)

실행:
- python excel_structure_extractor.py
"""

from __future__ import annotations

import json
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

# tkinter가 없는 환경에서도 즉시 크래시하지 않도록 보호
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception as _tk_error:  # noqa: N816
    tk = None
    filedialog = None
    messagebox = None
    ttk = None
    _TK_IMPORT_ERROR = _tk_error
else:
    _TK_IMPORT_ERROR = None

# xlwings는 선택 의존성
try:
    import xlwings as xw
except Exception:
    xw = None

APP_TITLE = "엑셀 구조 추출기"
DEFAULT_PREVIEW_ROWS = 10
DEFAULT_HEADER_SCAN_ROWS = 5
DEFAULT_MAX_COLS = 30


def safe_str(value: Any) -> str:
    return "" if value is None else str(value)


def mask_value(value: Any) -> str:
    text = safe_str(value)
    if not text:
        return ""

    text = re.sub(r"([A-Za-z0-9._%+-]+)@([A-Za-z0-9.-]+\.[A-Za-z]{2,})", "[EMAIL]", text)
    text = re.sub(r"\b[A-Z]{1,5}[-/]?\d{3,}\b", "[CODE]", text)
    text = re.sub(r"\b\d{4,}\b", "[NUM]", text)
    text = re.sub(r"(?<!\w)([\$₩¥€]?\s?\d{1,3}(?:,\d{3})+(?:\.\d+)?)", "[AMOUNT]", text)
    text = re.sub(r"(?<!\w)(\d+\.\d+)", "[NUMBER]", text)
    text = re.sub(r"\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b", "[DATE]", text)
    text = re.sub(r"\b\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}\b", "[DATE]", text)

    return text[:10] + "..." + text[-6:] if len(text) > 24 else text


def col_letter(n: int) -> str:
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result or "A"


def detect_non_empty_range(ws: Any, max_cols: int = DEFAULT_MAX_COLS) -> tuple[int, int]:
    max_row = ws.max_row or 1
    max_col = min(ws.max_column or 1, max_cols)

    last_non_empty_row = 1
    last_non_empty_col = 1

    for r in range(max_row, 0, -1):
        found = False
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if v is not None and safe_str(v).strip() != "":
                last_non_empty_row = r
                found = True
                break
        if found:
            break

    for c in range(max_col, 0, -1):
        found = False
        for r in range(1, min(max_row, 200) + 1):
            v = ws.cell(r, c).value
            if v is not None and safe_str(v).strip() != "":
                last_non_empty_col = c
                found = True
                break
        if found:
            break

    return last_non_empty_row, last_non_empty_col


def preview_sheet(ws: Any, preview_rows: int, max_cols: int, mask: bool) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    max_r = min(ws.max_row or 1, preview_rows)
    max_c = min(ws.max_column or 1, max_cols)

    for r in range(1, max_r + 1):
        row_data: dict[str, str] = {}
        for c in range(1, max_c + 1):
            cell_ref = f"{col_letter(c)}{r}"
            value = ws.cell(r, c).value
            text = safe_str(value)
            row_data[cell_ref] = mask_value(text) if mask else text
        rows.append(row_data)
    return rows


def find_formula_samples(
    ws: Any,
    max_rows: int = 200,
    max_cols: int = DEFAULT_MAX_COLS,
    limit: int = 20,
) -> list[str]:
    samples: list[str] = []
    mr = min(ws.max_row or 1, max_rows)
    mc = min(ws.max_column or 1, max_cols)

    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                samples.append(f"{col_letter(c)}{r}: {v}")
                if len(samples) >= limit:
                    return samples
    return samples


def summarize_workbook(
    file_path: str,
    preview_rows: int = DEFAULT_PREVIEW_ROWS,
    header_scan_rows: int = DEFAULT_HEADER_SCAN_ROWS,
    max_cols: int = DEFAULT_MAX_COLS,
    mask_preview: bool = True,
    include_formulas: bool = True,
) -> dict[str, Any]:
    # openpyxl import를 함수 내부로 옮겨 앱 시작 단계 크래시 방지
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl을 불러올 수 없습니다. 'pip install openpyxl' 후 다시 실행하세요."
        ) from exc

    wb = load_workbook(file_path, data_only=False, keep_vba=True)
    data_wb = load_workbook(file_path, data_only=True, keep_vba=True)

    result: dict[str, Any] = {
        "file_name": Path(file_path).name,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sheets": [],
        "defined_names": [],
    }

    try:
        try:
            dn = getattr(wb, "defined_names", None)
            if dn is not None and hasattr(dn, "definedName"):
                for item in dn.definedName:
                    result["defined_names"].append(
                        {
                            "name": getattr(item, "name", ""),
                            "value": getattr(item, "attr_text", "") or safe_str(item),
                        }
                    )
        except Exception:
            pass

        for ws, ws_data in zip(wb.worksheets, data_wb.worksheets):
            last_row, last_col = detect_non_empty_range(ws, max_cols=max_cols)

            header_rows: list[list[str]] = []
            for r in range(1, min(header_scan_rows, ws.max_row or 1) + 1):
                row_items: list[str] = []
                for c in range(1, min(ws.max_column or 1, max_cols) + 1):
                    v = ws.cell(r, c).value
                    text = safe_str(v)
                    row_items.append(f"{col_letter(c)}{r}={mask_value(text) if mask_preview else text}")
                header_rows.append(row_items)

            preview = preview_sheet(ws_data, preview_rows=preview_rows, max_cols=max_cols, mask=mask_preview)
            formulas = (
                find_formula_samples(ws, max_rows=200, max_cols=max_cols, limit=20)
                if include_formulas
                else []
            )

            result["sheets"].append(
                {
                    "sheet_name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "estimated_used_range": f"A1:{col_letter(last_col)}{last_row}",
                    "header_scan": header_rows,
                    "preview_rows": preview,
                    "formula_samples": formulas,
                }
            )
    finally:
        wb.close()
        data_wb.close()

    return result


def export_vba_modules_via_xlwings(file_path: str) -> dict[str, Any]:
    if xw is None:
        return {"success": False, "reason": "xlwings 미설치", "modules": []}

    app = None
    book = None
    modules: list[dict[str, Any]] = []

    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        book = app.books.open(file_path)
        api_book = book.api
        vbproject = api_book.VBProject
        vbcomponents = vbproject.VBComponents

        for i in range(1, vbcomponents.Count + 1):
            comp = vbcomponents.Item(i)
            code_module = comp.CodeModule
            line_count = code_module.CountOfLines
            code_text = code_module.Lines(1, line_count) if line_count > 0 else ""

            modules.append(
                {
                    "name": comp.Name,
                    "type": comp.Type,
                    "line_count": line_count,
                    "code": code_text,
                }
            )

        return {"success": True, "reason": "", "modules": modules}
    except Exception as exc:
        return {"success": False, "reason": f"{type(exc).__name__}: {exc}", "modules": []}
    finally:
        try:
            if book is not None:
                book.close()
        except Exception:
            pass
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass


def render_report_text(summary: dict[str, Any], vba_info: Optional[dict[str, Any]]) -> str:
    lines: list[str] = []
    lines.append("# 엑셀 구조 추출 보고서")
    lines.append("")
    lines.append(f"- 파일명: {summary.get('file_name', '')}")
    lines.append(f"- 생성시각: {summary.get('generated_at', '')}")
    lines.append("")

    defined_names = summary.get("defined_names", [])
    if defined_names:
        lines.append("## 이름 정의")
        for item in defined_names:
            lines.append(f"- {item.get('name', '')}: {item.get('value', '')}")
        lines.append("")

    lines.append("## 시트 요약")
    for sheet in summary.get("sheets", []):
        lines.append(f"### {sheet['sheet_name']}")
        lines.append(f"- 크기: {sheet['max_row']}행 x {sheet['max_column']}열")
        lines.append(f"- 추정 사용범위: {sheet['estimated_used_range']}")
        lines.append("")

        lines.append("#### 헤더 스캔")
        for row in sheet["header_scan"]:
            lines.append("- " + " | ".join(row))
        lines.append("")

        lines.append("#### 미리보기")
        for row in sheet["preview_rows"]:
            parts = [f"{k}={v}" for k, v in row.items()]
            lines.append("- " + " | ".join(parts))
        lines.append("")

        if sheet["formula_samples"]:
            lines.append("#### 수식 샘플")
            for formula in sheet["formula_samples"]:
                lines.append(f"- {formula}")
            lines.append("")

    if vba_info is not None:
        lines.append("## VBA 모듈")
        if vba_info.get("success"):
            for module in vba_info.get("modules", []):
                lines.append(
                    f"### 모듈명: {module['name']} / 타입: {module['type']} / 줄수: {module['line_count']}"
                )
                lines.append("```vb")
                lines.append(module["code"][:50000])
                lines.append("```")
                lines.append("")
        else:
            lines.append(f"- VBA 추출 실패: {vba_info.get('reason', '')}")
            lines.append("- xlwings 미설치/미지원 환경이어도 시트 구조 추출은 계속 진행됩니다.")
            lines.append("- Windows + Excel 설치 + 보안 설정(신뢰할 수 있는 VBA 접근) 확인 권장")
            lines.append("")

    return "\n".join(lines)


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("920x760")
        self.root.minsize(840, 680)

        self.file_path = tk.StringVar()
        self.preview_rows = tk.StringVar(value=str(DEFAULT_PREVIEW_ROWS))
        self.header_scan_rows = tk.StringVar(value=str(DEFAULT_HEADER_SCAN_ROWS))
        self.max_cols = tk.StringVar(value=str(DEFAULT_MAX_COLS))
        self.mask_preview = tk.BooleanVar(value=True)
        self.include_formulas = tk.BooleanVar(value=True)
        self.include_vba = tk.BooleanVar(value=True)

        self._build_ui()
        self._show_startup_notice()

    def _build_ui(self) -> None:
        outer = ttk.Frame(self.root, padding=14)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text="엑셀 구조 추출기", font=("Malgun Gothic", 16, "bold")).pack(anchor="w")
        ttk.Label(
            outer,
            text="원본 파일을 읽어 시트 구조/헤더/수식/VBA(선택)를 보고서로 저장합니다.",
        ).pack(anchor="w", pady=(4, 12))

        file_box = ttk.LabelFrame(outer, text="1) 파일 선택", padding=10)
        file_box.pack(fill="x", pady=6)

        ttk.Entry(file_box, textvariable=self.file_path).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(file_box, text="엑셀 파일 열기", command=self.choose_file).grid(row=0, column=1)
        file_box.columnconfigure(0, weight=1)

        option_box = ttk.LabelFrame(outer, text="2) 추출 옵션", padding=10)
        option_box.pack(fill="x", pady=6)

        ttk.Label(option_box, text="미리보기 행 수").grid(row=0, column=0, sticky="w")
        ttk.Entry(option_box, textvariable=self.preview_rows, width=10).grid(row=0, column=1, sticky="w", padx=(6, 14))
        ttk.Label(option_box, text="헤더 스캔 행 수").grid(row=0, column=2, sticky="w")
        ttk.Entry(option_box, textvariable=self.header_scan_rows, width=10).grid(row=0, column=3, sticky="w", padx=(6, 14))
        ttk.Label(option_box, text="최대 열 수").grid(row=0, column=4, sticky="w")
        ttk.Entry(option_box, textvariable=self.max_cols, width=10).grid(row=0, column=5, sticky="w", padx=(6, 0))

        ttk.Checkbutton(option_box, text="미리보기 값 마스킹", variable=self.mask_preview).grid(
            row=1, column=0, sticky="w", pady=(10, 0)
        )
        ttk.Checkbutton(option_box, text="수식 샘플 포함", variable=self.include_formulas).grid(
            row=1, column=1, columnspan=2, sticky="w", pady=(10, 0)
        )
        ttk.Checkbutton(option_box, text="VBA 모듈 추출 시도", variable=self.include_vba).grid(
            row=1, column=3, columnspan=3, sticky="w", pady=(10, 0)
        )

        btn_box = ttk.Frame(outer)
        btn_box.pack(fill="x", pady=(10, 8))
        ttk.Button(btn_box, text="보고서 생성", command=self.run_extract).pack(side="left")
        ttk.Button(btn_box, text="로그 지우기", command=self.clear_log).pack(side="left", padx=(8, 0))

        log_box = ttk.LabelFrame(outer, text="3) 진행 로그", padding=10)
        log_box.pack(fill="both", expand=True, pady=6)
        self.log_text = tk.Text(log_box, height=24, font=("Consolas", 10), wrap="word")
        self.log_text.pack(fill="both", expand=True)

    def _show_startup_notice(self) -> None:
        self.log("프로그램 시작 완료")
        self.log("필수: openpyxl / 선택: xlwings")
        if xw is None:
            self.log("[경고] xlwings를 찾지 못했습니다. VBA 추출 없이 계속 진행합니다.")

    def log(self, text: str) -> None:
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def clear_log(self) -> None:
        self.log_text.delete("1.0", "end")

    def choose_file(self) -> None:
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if path:
            self.file_path.set(path)
            self.log(f"파일 선택: {path}")

    def run_extract(self) -> None:
        try:
            file_path = self.file_path.get().strip()
            if not file_path:
                raise ValueError("엑셀 파일을 선택하세요.")
            if not Path(file_path).exists():
                raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

            preview_rows = int(self.preview_rows.get().strip())
            header_scan_rows = int(self.header_scan_rows.get().strip())
            max_cols = int(self.max_cols.get().strip())
            if preview_rows <= 0 or header_scan_rows <= 0 or max_cols <= 0:
                raise ValueError("옵션 값은 1 이상의 정수여야 합니다.")

            self.log("시트 구조 추출 시작...")
            summary = summarize_workbook(
                file_path=file_path,
                preview_rows=preview_rows,
                header_scan_rows=header_scan_rows,
                max_cols=max_cols,
                mask_preview=self.mask_preview.get(),
                include_formulas=self.include_formulas.get(),
            )
            self.log("시트 구조 추출 완료")

            vba_info = None
            if self.include_vba.get():
                self.log("VBA 모듈 추출 시도...")
                vba_info = export_vba_modules_via_xlwings(file_path)
                if vba_info.get("success"):
                    self.log(f"VBA 추출 완료: {len(vba_info.get('modules', []))}개 모듈")
                else:
                    self.log(f"[경고] VBA 추출 실패: {vba_info.get('reason', '')}")

            report_text = render_report_text(summary, vba_info)

            out_dir = Path(file_path).parent
            stem = Path(file_path).stem
            txt_path = out_dir / f"{stem}_structure_report.txt"
            md_path = out_dir / f"{stem}_structure_report.md"
            json_path = out_dir / f"{stem}_structure_report.json"

            txt_path.write_text(report_text, encoding="utf-8")
            md_path.write_text(report_text, encoding="utf-8")
            json_path.write_text(
                json.dumps({"summary": summary, "vba_info": vba_info}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            self.log(f"저장 완료: {txt_path}")
            self.log(f"저장 완료: {md_path}")
            self.log(f"저장 완료: {json_path}")
            messagebox.showinfo(APP_TITLE, f"보고서 생성 완료\n\nTXT: {txt_path}\nMD: {md_path}\nJSON: {json_path}")
        except Exception as exc:
            self.log("[오류]")
            self.log(str(exc))
            self.log(traceback.format_exc())
            try:
                messagebox.showerror(APP_TITLE, f"오류 발생\n\n{exc}")
            except Exception:
                pass


def main() -> int:
    # 1) tkinter 문제를 친절히 안내하고 종료
    if tk is None or ttk is None or filedialog is None or messagebox is None:
        print("[치명적] tkinter를 불러올 수 없어 GUI를 시작할 수 없습니다.")
        print(f"원인: {_TK_IMPORT_ERROR}")
        print("해결: Windows 기본 Python 또는 tkinter 포함된 Python 배포판을 사용하세요.")
        return 1

    # 2) 루트 생성 단계 오류 보호 (디스플레이/권한/환경 문제 등)
    try:
        root = tk.Tk()
    except Exception as exc:
        print("[치명적] GUI 창 생성 실패")
        print(f"원인: {type(exc).__name__}: {exc}")
        return 1

    try:
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        App(root)
        root.mainloop()
        return 0
    except Exception as exc:
        print("[치명적] 프로그램 초기화 실패")
        print(f"원인: {type(exc).__name__}: {exc}")
        print(traceback.format_exc())
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
